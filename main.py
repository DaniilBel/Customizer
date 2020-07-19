from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as opx
from tqdm.auto import tqdm


class PriceConverter:
    def __init__(self, data):
        self.site = None
        self.data = data
        self.reformatDate = []
        self.dollarRates = []

        # Костыль
        self.tmpDollar = ""
        self.tds = []

    def pullReformatDate(self):
        for date, _ in self.data.items():
            self.reformatDate.append(date[0].value)

    # Получаем сайт с нужной датой
    def parseDate(self, i):
        siteNameL = "https://www.cbr.ru/currency_base/daily/?UniDbQuery.Posted=True&UniDbQuery.To="

        # dd/mm/yyyy
        d = self.reformatDate[i].split("/")

        self.site = siteNameL + d[0] + "." + d[1] + "." + d[2]

    def parseCourse(self, dollar):
        return dollar.replace(",", ".")

    # Получаем курс доллара на текущую дату
    # В случае совпадения дат, не выполняем эту функцию
    def pickCourse(self, soup):
        # Костыль
        for tag in soup.find_all('td'):
            self.tds.append(tag.text)

        dollar = self.tds[self.tds.index("Доллар США") + 1]
        self.tmpDollar = dollar

        return self.parseCourse(dollar)

    def pickData(self):
        self.pullReformatDate()
        tmp = ""

        for i in tqdm(range(len(self.reformatDate)), desc="Num of parsed site"):
            if tmp == self.reformatDate[i]:
                self.dollarRates.append(self.parseCourse(self.tmpDollar))
                continue
            else:
                self.parseDate(i)
                html = urlopen(self.site).read().decode('utf-8')
                soup = BeautifulSoup(html, 'html.parser')

                # Список курсов доллара для каждого нужного дня
                self.dollarRates.append(self.pickCourse(soup))
                self.tds = []
                tmp = self.reformatDate[i]

        return self.dollarRates


class ExcelTable:
    def __init__(self, filename):
        self.wb = None  # Workbook
        self.filename = filename
        self.sheet = None  # Таблица
        self.date_comm = {}

    def read(self):
        self.wb = opx.load_workbook(filename=self.filename)
        self.sheet = self.wb.active
        return self

    def reformatComm(self):
        for _, comm in self.date_comm.items():
            v = comm[0].value
            comm[0].value = v.replace(",", ".")

        return self.date_comm

    def pickData(self):
        # дата и комиссия в эту дату
        self.date_comm = dict(zip(self.sheet.iter_rows(min_row=9, max_row=self.sheet.max_row - 1, min_col=1, max_col=1),
                                  self.sheet.iter_rows(min_row=9, max_row=self.sheet.max_row - 1, min_col=8,
                                                       max_col=8)))

        # Возвращаем отформатированные данные (дату и комиссию)
        return self.reformatComm()

    # Записываем в ячейки результат
    def write(self, data):
        i = 9
        for rec in data:
            self.sheet.cell(row=i, column=10).value = rec
            i += 1

    # "filename.xlsx"
    def save(self):
        self.wb.save(self.filename[:-5] + "Done" + self.filename[-5:])


# Комиссию сопоставляем с курсом доллара на определенную дату
def arrange(data, course):
    # print(len(data), len(course))
    res = {}
    # Надо будет переделать
    if len(data) == len(course):
        i = 0
        for _, comm in data.items():
            res.update({(float(comm[0].value), i): float(course[i])})
            i += 1
        return res
    else:
        return 0


# Вычисляем комиссию в рублях, исходя из текущего курса
# Комиссия дана в долларах
def evaluate(data):
    res = []
    for key, value in data.items():
        res.append(round(key[0] * value, 6))

    return res


if __name__ == '__main__':
    et = ExcelTable("PAypal.xlsx").read()
    date_comm = et.pickData()

    # for date, comm in d.items():
    #     print(date[0].value, comm[0].value, sep=' ')

    pc = PriceConverter(date_comm)
    course = pc.pickData()

    # print(course)

    comm_course = arrange(date_comm, course)
    commRub = evaluate(comm_course)

    et.write(commRub)
    et.save()
